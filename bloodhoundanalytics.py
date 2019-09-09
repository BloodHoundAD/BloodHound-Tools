from openpyxl import Workbook, styles
from openpyxl.utils import get_column_letter
from neo4j.v1 import GraphDatabase
import cmd
import sys
from timeit import default_timer as timer

# Authors:
# Andy Robbins - @_wald0
# Rohan Vazarkar - @CptJesus
# https://www.specterops.io

# License: GPLv3


class Messages(object):
    def title(self):
        print "================================================================"
        print " BloodHound Analytics Generator"
        print "   connect - connect to database"
        print "   dbconfig - configure database settings"
        print "   changedomain - change domain for analysis"
        print "   startanalysis - start analysis for specified domain"
        print "   changefilename - change output filename"
        print "================================================================"

    def input_default(self, prompt, default):
        return raw_input("%s [%s] " % (prompt, default)) or default

class FrontPage(object):
    def __init__(self, driver, domain, workbook):
        self.driver = driver
        self.domain = domain
        self.col_count = 1
        self.workbook = workbook

    # Modified, dont copy this
    def write_column_data(self, sheet, title, results):
        count = len(results)
        offset = 1
        count_cell = 4
        font = styles.Font(bold=True)
        c = sheet.cell(offset, count_cell)
        c.font = font
        sheet.cell(offset, count_cell, value=title.format(count))
        for i in xrange(0, count):
            sheet.cell(i + offset + 1, count_cell, value=results[i])
        count_cell += 1

    def write_single_cell(self, sheet, row, column, text):
        sheet.cell(row, column, value=text)

    def do_front_page_analysis(self):
        func_list = [self.create_node_statistics, self.create_highvalue_list,
                     self.create_edge_statistics, self.create_qa_statistics,
                     ]
        sheet = self.workbook._sheets[0]
        self.write_single_cell(sheet, 1, 1, "Node Statistics")
        self.write_single_cell(sheet, 1, 2, "Edge Statistics")
        self.write_single_cell(sheet, 1, 3, "QA Info")
        font = styles.Font(bold=True)
        sheet.cell(1, 1).font = font
        sheet.cell(1, 2).font = font
        sheet.cell(1, 3).font = font

        for f in func_list:
            s = timer()
            f(sheet)
            print "{} completed in {}s".format(f.__name__, timer() - s)

    def create_node_statistics(self, sheet):
        session = self.driver.session()
        for result in session.run("MATCH (n:User {domain:{domain}}) RETURN count(n)", domain=self.domain):
            self.write_single_cell(
                sheet, 2, 1, "Users: {:,}".format(result[0]))

        for result in session.run("MATCH (n:Group {domain:{domain}}) RETURN count(n)", domain=self.domain):
            self.write_single_cell(
                sheet, 3, 1, "Groups: {:,}".format(result[0]))

        for result in session.run("MATCH (n:Computer {domain:{domain}}) RETURN count(n)", domain=self.domain):
            self.write_single_cell(
                sheet, 4, 1, "Computers: {:,}".format(result[0]))

        for result in session.run("MATCH (n:Domain) RETURN count(n)", domain=self.domain):
            self.write_single_cell(
                sheet, 5, 1, "Other Domains: {:,}".format(result[0] - 1))

        for result in session.run("MATCH (n:GPO) WHERE n.name =~ '.*" + self.domain + "$' RETURN count(n)"):
            self.write_single_cell(
                sheet, 6, 1, "GPOs: {:,}".format(result[0]))

        for result in session.run("MATCH (n:OU) WHERE n.name =~ '.*@" + self.domain + "$' RETURN count(n)"):
            self.write_single_cell(
                sheet, 7, 1, "OUs: {:,}".format(result[0]))

        session.close()

    def create_edge_statistics(self, sheet):
        session = self.driver.session()
        for result in session.run("MATCH ()-[r:MemberOf]->({domain:{domain}}) RETURN count(r)", domain=self.domain):
            self.write_single_cell(
                sheet, 2, 2, "MemberOf: {:,}".format(result[0]))

        for result in session.run("MATCH ()-[r:AdminTo]->({domain:{domain}}) RETURN count(r)", domain=self.domain):
            self.write_single_cell(
                sheet, 3, 2, "AdminTo: {:,}".format(result[0]))

        for result in session.run("MATCH ()-[r:HasSession]->({domain:{domain}}) RETURN count(r)", domain=self.domain):
            self.write_single_cell(
                sheet, 4, 2, "HasSession: {:,}".format(result[0]))

        for result in session.run("MATCH ()-[r:GpLink]-(n) WHERE n.name =~ '.*" + self.domain + "$' RETURN count(r)"):
            self.write_single_cell(
                sheet, 5, 2, "GpLinks: {:,}".format(result[0]))

        for result in session.run("MATCH ()-[r {isacl:true}]->({domain:{domain}}) RETURN count(r)", domain=self.domain):
            self.write_single_cell(
                sheet, 6, 2, "ACLs: {:,}".format(result[0]))
        session.close()

    def create_qa_statistics(self, sheet):
        session = self.driver.session()
        computer_local_admin_pct = 0
        computer_session_pct = 0
        user_session_pct = 0

        query = """MATCH (n)-[:AdminTo]->(c:Computer {domain:{domain}})
                    WITH COUNT(DISTINCT(c)) as computersWithAdminsCount
                    MATCH (c2:Computer {domain:{domain}})
                    RETURN toInt(100 * (toFloat(computersWithAdminsCount) / COUNT(c2)))
                    """
        for result in session.run(query, domain=self.domain):
            computer_local_admin_pct = result[0]

        query = """MATCH (c:Computer {domain:{domain}})-[:HasSession]->()
                    WITH COUNT(DISTINCT(c)) as computersWithSessions
                    MATCH (c2:Computer {domain:{domain}})
                    RETURN toInt(100 * (toFloat(computersWithSessions) / COUNT(c2)))
                    """

        for result in session.run(query, domain=self.domain):
            computer_session_pct = result[0]

        query = """MATCH ()-[:HasSession]->(u:User {domain:{domain}})
                    WITH COUNT(DISTINCT(u)) as usersWithSessions
                    MATCH (u2:User {domain:{domain},enabled:true})
                    RETURN toInt(100 * (toFloat(usersWithSessions) / COUNT(u2)))
                    """

        for result in session.run(query, domain=self.domain):
            user_session_pct = result[0]

        query = """MATCH (u:User {domain: {domain}})
                    MATCH (g:Group {domain: {domain}})
                    WHERE g.objectsid ENDS WITH '-512'
                    WITH g, COUNT(u) as userCount
                    MATCH p = shortestPath((u:User {domain: {domain}})-[*1..]->(g))
                    RETURN toint(100.0 * COUNT(u) / userCount)
                    """

        for result in session.run(query, domain=self.domain):
            Users_to_da = result[0]

        query = """MATCH (c:Computer {domain: {domain}})
                    MATCH (g:Group {domain: {domain}})
                    WHERE g.objectsid ENDS WITH '-512'
                    WITH g, COUNT(c) as ComputerCount
                    MATCH p = shortestPath((c:Computer {domain: {domain}})-[*1..]->(g))
                    RETURN toint(100.0 * COUNT(c) / ComputerCount)
                    """

        for result in session.run(query, domain=self.domain):
            Computers_to_da = result[0]

        query = """MATCH(n:User)
                    WHERE n.name =~ 'GUEST@.*'
                    RETURN n.enabled
                    """

        for result in session.run(query, domain=self.domain):
            Guest_account = result[0]

        session.close()
        self.write_single_cell(sheet, 2, 3, "Computers With Local Admin Data: {}%".format(
            computer_local_admin_pct))
        self.write_single_cell(
            sheet, 3, 3, "Computers With Session Data: {}%".format(computer_session_pct))
        self.write_single_cell(
            sheet, 4, 3, "Users With Session Data: {}%".format(user_session_pct))
        self.write_single_cell(
            sheet, 5, 3, "Users with attack path to Domain Admin: {}%".format(Users_to_da))
        self.write_single_cell(
            sheet, 6, 3, "Computers with attack path to Domain Admin: {}%".format(Computers_to_da))
        self.write_single_cell(
            sheet, 7, 3, "Guest Account: {}".format(Guest_account))

    def create_highvalue_list(self, sheet):
        list_query = """MATCH (n {highvalue: True,domain: {domain}})
                            RETURN DISTINCT n.name
                            ORDER BY n.name ASC
                            """
        session = self.driver.session()
        results = []

        for result in session.run(list_query, domain=self.domain):
            results.append(result[0])

        session.close()
        self.write_column_data(sheet, "High Value List: {}", results)

class LowHangingFruit(object):
    def __init__(self, driver, domain, workbook):
        self.driver = driver
        self.domain = domain
        self.col_count = 1
        self.workbook = workbook

    def write_column_data(self, sheet, title, results):
        count = len(results)
        offset = 6
        font = styles.Font(bold=True)
        c = sheet.cell(offset, self.col_count)
        c.font = font
        sheet.cell(offset, self.col_count, value=title.format(count))
        for i in xrange(0, count):
            sheet.cell(i + offset + 1, self.col_count, value=results[i])
        self.col_count += 1

    def write_single_cell(self, sheet, row, column, text):
        sheet.cell(row, column, value=text)

    def do_low_hanging_fruit_analysis(self):
        func_list = [
            self.domain_user_admin, self.everyone_admin, self.authenticated_users_admin,
            self.domain_users_control, self.everyone_control, self.authenticated_users_control,
            self.domain_users_rdp, self.everyone_rdp, self.authenticated_users_dcom,
            self.domain_users_dcom, self.everyone_dcom, self.authenticated_users_dcom,
            self.shortest_acl_path_domain_users, self.shortest_derivative_path_domain_users,
            self.shortest_hybrid_path_domain_users, self.shortest_acl_path_everyone,
            self.shortest_derivative_path_everyone, self.shortest_hybrid_path_everyone,
            self.shortest_acl_path_auth_users, self.shortest_derivative_path_auth_users,
            self.shortest_hybrid_path_auth_users, self.kerberoastable_path_len,
            self.asreproastable_path_len, self.high_admin_comps, self.server_2003,
            self.server_2008, self.spoolbug, self.passnotreqd
        ]
        sheet = self.workbook._sheets[2]
        self.write_single_cell(sheet, 1, 1, "Domain Users to Domain Admins")
        self.write_single_cell(sheet, 1, 2, "Everyone to Domain Admins")
        self.write_single_cell(sheet, 1, 3, "Authenticated Users to Domain Admins")
        font = styles.Font(bold=True)
        sheet.cell(1, 1).font = font
        sheet.cell(1, 2).font = font
        sheet.cell(1, 3).font = font

        for f in func_list:
            s = timer()
            f(sheet)
            print "{} completed in {}s".format(f.__name__, timer() - s)

    def domain_user_admin(self, sheet):
        list_query = """MATCH (g:Group {domain:{domain}})
                    WHERE g.objectsid ENDS WITH "-513"
                    OPTIONAL MATCH (g)-[:AdminTo]->(c1)
                    OPTIONAL MATCH (g)-[:MemberOf*1..]->(:Group)-[:AdminTo]->(c2)
                    WITH COLLECT(c1) + COLLECT(c2) as tempVar
                    UNWIND tempVar AS computers
                    RETURN DISTINCT(computers.name)
                    """

        session = self.driver.session()
        results = []

        for result in session.run(list_query, domain=self.domain):
            results.append(result[0])

        session.close()
        self.write_column_data(
            sheet, "Domain Users with Local Admin: {}", results)

    def everyone_admin(self, sheet):
        list_query = """MATCH (g:Group {domain:{domain}})
                        WHERE g.objectsid = "S-1-1-0"
                        OPTIONAL MATCH (g)-[:AdminTo]->(c1)
                        OPTIONAL MATCH (g)-[:MemberOf*1..]->(:Group)-[:AdminTo]->(c2)
                        WITH COLLECT(c1) + COLLECT(c2) as tempVar
                        UNWIND tempVar AS computers
                        RETURN DISTINCT(computers.name)
                        """

        session = self.driver.session()
        results = []

        for result in session.run(list_query, domain=self.domain):
            results.append(result[0])

        session.close()
        self.write_column_data(sheet, "Everyone with Local Admin: {}", results)

    def authenticated_users_admin(self, sheet):

        list_query = """MATCH (g:Group {domain:{domain}})
                        WHERE g.objectsid = "S-1-5-11"
                        OPTIONAL MATCH (g)-[:AdminTo]->(c1)
                        OPTIONAL MATCH (g)-[:MemberOf*1..]->(:Group)-[:AdminTo]->(c2)
                        WITH COLLECT(c1) + COLLECT(c2) as tempVar
                        UNWIND tempVar AS computers
                        RETURN DISTINCT(computers.name)
                        """

        session = self.driver.session()
        results = []

        for result in session.run(list_query, domain=self.domain):
            results.append(result[0])

        session.close()
        self.write_column_data(
            sheet, "Authenticated Users with Local Admin: {}", results)

    def domain_users_control(self, sheet):

        list_query = """MATCH (g:Group {domain:{domain}})
                        WHERE g.objectsid ENDS WITH "-513"
                        OPTIONAL MATCH (g)-[{isacl:true}]->(n)
                        OPTIONAL MATCH (g)-[:MemberOf*1..]->(:Group)-[{isacl:true}]->(m)
                        WITH COLLECT(n) + COLLECT(m) as tempVar
                        UNWIND tempVar AS objects
                        RETURN DISTINCT(objects)
                        ORDER BY objects.name ASC
                        """

        session = self.driver.session()
        results = []

        for result in session.run(list_query, domain=self.domain):
            results.append(result[0])

        session.close()
        self.write_column_data(
            sheet, "Objects Controlled by Domain Users: {}", results)

    def everyone_control(self, sheet):

        list_query = """MATCH (g:Group {domain:{domain}})
                        WHERE g.objectsid = 'S-1-1-0'
                        OPTIONAL MATCH (g)-[{isacl:true}]->(n)
                        OPTIONAL MATCH (g)-[:MemberOf*1..]->(:Group)-[{isacl:true}]->(m)
                        WITH COLLECT(n) + COLLECT(m) as tempVar
                        UNWIND tempVar AS objects
                        RETURN DISTINCT(objects)
                        ORDER BY objects.name ASC
                        """

        session = self.driver.session()
        results = []

        for result in session.run(list_query, domain=self.domain):
            results.append(result[0])

        session.close()
        self.write_column_data(
            sheet, "Objects Controlled by Everyone: {}", results)

    def authenticated_users_control(self, sheet):

        list_query = """MATCH (g:Group {domain:{domain}})
                        WHERE g.objectsid = 'S-1-5-11'
                        OPTIONAL MATCH (g)-[{isacl:true}]->(n)
                        OPTIONAL MATCH (g)-[:MemberOf*1..]->(:Group)-[{isacl:true}]->(m)
                        WITH COLLECT(n) + COLLECT(m) as tempVar
                        UNWIND tempVar AS objects
                        RETURN DISTINCT(objects)
                        ORDER BY objects.name ASC
                        """

        session = self.driver.session()
        results = []

        for result in session.run(list_query, domain=self.domain):
            results.append(result[0])

        session.close()
        self.write_column_data(
            sheet, "Objects Controlled by Authenticated Users: {}", results)

    def domain_users_rdp(self, sheet):

        list_query = """MATCH (g:Group {domain:{domain}})
                        WHERE g.objectsid ENDS WITH "-513"
                        OPTIONAL MATCH (g)-[:CanRDP]->(c1)
                        OPTIONAL MATCH (g)-[:MemberOf*1..]->(:Group)-[:CanRDP]->(c2)
                        WITH COLLECT(c1) + COLLECT(c2) as tempVar
                        UNWIND tempVar AS computers
                        RETURN DISTINCT(computers.name)
                        """

        session = self.driver.session()
        results = []

        for result in session.run(list_query, domain=self.domain):
            results.append(result[0])

        session.close()
        self.write_column_data(
            sheet, "Domain Users with RDP Rights: {}", results)

    def everyone_rdp(self, sheet):

        list_query = """MATCH (g:Group {domain:{domain}})
                        WHERE g.objectsid = "S-1-1-0"
                        OPTIONAL MATCH (g)-[:CanRDP]->(c1)
                        OPTIONAL MATCH (g)-[:MemberOf*1..]->(:Group)-[:CanRDP]->(c2)
                        WITH COLLECT(c1) + COLLECT(c2) as tempVar
                        UNWIND tempVar AS computers
                        RETURN DISTINCT(computers.name)
                        """

        session = self.driver.session()
        results = []

        for result in session.run(list_query, domain=self.domain):
            results.append(result[0])

        session.close()
        self.write_column_data(sheet, "Everyone with RDP Rights: {}", results)

    def authenticated_users_rdp(self, sheet):

        list_query = """MATCH (g:Group {domain:{domain}})
                        WHERE g.objectsid = "S-1-5-11"
                        OPTIONAL MATCH (g)-[:CanRDP]->(c1)
                        OPTIONAL MATCH (g)-[:MemberOf*1..]->(:Group)-[:CanRDP]->(c2)
                        WITH COLLECT(c1) + COLLECT(c2) as tempVar
                        UNWIND tempVar AS computers
                        RETURN DISTINCT(computers.name)
                        """

        session = self.driver.session()
        results = []

        for result in session.run(list_query, domain=self.domain):
            results.append(result[0])

        session.close()
        self.write_column_data(
            sheet, "Authenticated Users with RDP Rights: {}", results)

    def domain_users_dcom(self, sheet):

        list_query = """MATCH (g:Group {domain:{domain}})
                        WHERE g.objectsid ENDS WITH "-513"
                        OPTIONAL MATCH (g)-[:ExecuteDCOM]->(c1)
                        OPTIONAL MATCH (g)-[:MemberOf*1..]->(:Group)-[:ExecuteDCOM]->(c2)
                        WITH COLLECT(c1) + COLLECT(c2) as tempVar
                        UNWIND tempVar AS computers
                        RETURN DISTINCT(computers.name)
                        """

        session = self.driver.session()
        results = []

        for result in session.run(list_query, domain=self.domain):
            results.append(result[0])

        session.close()
        self.write_column_data(
            sheet, "Domain Users with DCOM Rights: {}", results)

    def everyone_dcom(self, sheet):

        list_query = """MATCH (g:Group {domain:{domain}})
                        WHERE g.objectsid = "S-1-1-0"
                        OPTIONAL MATCH (g)-[:ExecuteDCOM]->(c1)
                        OPTIONAL MATCH (g)-[:MemberOf*1..]->(:Group)-[:ExecuteDCOM]->(c2)
                        WITH COLLECT(c1) + COLLECT(c2) as tempVar
                        UNWIND tempVar AS computers
                        RETURN DISTINCT(computers.name)
                        """

        session = self.driver.session()
        results = []

        for result in session.run(list_query, domain=self.domain):
            results.append(result[0])

        session.close()
        self.write_column_data(
            sheet, "Domain Users with DCOM Rights: {}", results)

    def authenticated_users_dcom(self, sheet):

        list_query = """MATCH (g:Group {domain:{domain}})
                        WHERE g.objectsid = "S-1-5-11"
                        OPTIONAL MATCH (g)-[:ExecuteDCOM]->(c1)
                        OPTIONAL MATCH (g)-[:MemberOf*1..]->(:Group)-[:ExecuteDCOM]->(c2)
                        WITH COLLECT(c1) + COLLECT(c2) as tempVar
                        UNWIND tempVar AS computers
                        RETURN DISTINCT(computers.name)
                        """

        session = self.driver.session()
        results = []

        for result in session.run(list_query, domain=self.domain):
            results.append(result[0])

        session.close()
        self.write_column_data(
            sheet, "Domain Users with DCOM Rights: {}", results)

    def shortest_acl_path_domain_users(self, sheet):
        count_query = """MATCH (g1:Group {domain:{domain}})
                        WHERE g1.objectsid ENDS WITH "-513"
                        MATCH (g2:Group {domain:{domain}})
                        WHERE g2.objectsid ENDS WITH "-512"
                        MATCH p = shortestPath((g1)-[:Owns|AllExtendedRights|ForceChangePassword|GenericAll|GenericWrite|WriteDacl|WriteOwner*1..]->(g2))
                        RETURN LENGTH(p)
                        """

        session = self.driver.session()
        count = 0
        for result in session.run(count_query, domain=self.domain):
            count = result[0]

        session.close()
        self.write_single_cell(
            sheet, 2, 1, "Shortest ACL Path Length: {}".format(count))

    def shortest_derivative_path_domain_users(self, sheet):
        count_query = """MATCH (g1:Group {domain:{domain}})
                        WHERE g1.objectsid ENDS WITH "-513"
                        MATCH (g2:Group {domain:{domain}})
                        WHERE g2.objectsid ENDS WITH "-512"
                        MATCH p = shortestPath((g1)-[:AdminTo|HasSession|MemberOf*1..]->(g2))
                        RETURN LENGTH(p)
                        """

        session = self.driver.session()
        count = 0
        for result in session.run(count_query, domain=self.domain):
            count = result[0]

        session.close()
        self.write_single_cell(
            sheet, 3, 1, "Shortest Derivative Path Length: {}".format(count))

    def shortest_hybrid_path_domain_users(self, sheet):
        count_query = """MATCH (g1:Group {domain:{domain}})
                        WHERE g1.objectsid ENDS WITH "-513"
                        MATCH (g2:Group {domain:{domain}})
                        WHERE g2.objectsid ENDS WITH "-512"
                        MATCH p = shortestPath((g1)-[r*1..]->(g2))
                        WHERE NONE(rel in r WHERE type(rel)="GetChanges")
                        WITH *
                        WHERE NONE(rel in r WHERE type(rel)="GetChangesAll")
                        RETURN LENGTH(p)
                        """

        session = self.driver.session()
        count = 0
        for result in session.run(count_query, domain=self.domain):
            count = result[0]

        session.close()
        self.write_single_cell(
            sheet, 4, 1, "Shortest Hybrid Path Length: {}".format(count))

    def shortest_acl_path_everyone(self, sheet):
        count_query = """MATCH (g1:Group {domain:{domain}})
                        WHERE g1.objectsid = 'S-1-1-0'
                        MATCH (g2:Group {domain:{domain}})
                        WHERE g2.objectsid ENDS WITH "-512"
                        MATCH p = shortestPath((g1)-[:Owns|AllExtendedRights|ForceChangePassword|GenericAll|GenericWrite|WriteDacl|WriteOwner*1..]->(g2))
                        RETURN LENGTH(p)
                        """

        session = self.driver.session()
        count = 0
        for result in session.run(count_query, domain=self.domain):
            count = result[0]

        session.close()
        self.write_single_cell(
            sheet, 2, 2, "Shortest ACL Path Length: {}".format(count))

    def shortest_derivative_path_everyone(self, sheet):
        count_query = """MATCH (g1:Group {domain:{domain}})
                        WHERE g1.objectsid = 'S-1-1-0'
                        MATCH (g2:Group {domain:{domain}})
                        WHERE g2.objectsid ENDS WITH "-512"
                        MATCH p = shortestPath((g1)-[:AdminTo|HasSession|MemberOf*1..]->(g2))
                        RETURN LENGTH(p)
                        """

        session = self.driver.session()
        count = 0
        for result in session.run(count_query, domain=self.domain):
            count = result[0]

        session.close()
        self.write_single_cell(
            sheet, 3, 2, "Shortest Derivative Path Length: {}".format(count))

    def shortest_hybrid_path_everyone(self, sheet):
        count_query = """MATCH (g1:Group {domain:{domain}})
                        WHERE g1.objectsid = 'S-1-1-0'
                        MATCH (g2:Group {domain:{domain}})
                        WHERE g2.objectsid ENDS WITH "-512"
                        MATCH p = shortestPath((g1)-[r*1..]->(g2))
                        WHERE NONE(rel in r WHERE type(rel)="GetChanges")
                        WITH *
                        WHERE NONE(rel in r WHERE type(rel)="GetChangesAll")
                        RETURN LENGTH(p)
                        """

        session = self.driver.session()
        count = 0
        for result in session.run(count_query, domain=self.domain):
            count = result[0]

        session.close()
        self.write_single_cell(
            sheet, 4, 2, "Shortest Hybrid Path Length: {}".format(count))

    def shortest_acl_path_auth_users(self, sheet):
        count_query = """MATCH (g1:Group {domain:{domain}})
                        WHERE g1.objectsid = 'S-1-5-11'
                        MATCH (g2:Group {domain:{domain}})
                        WHERE g2.objectsid ENDS WITH "-512"
                        MATCH p = shortestPath((g1)-[:Owns|AllExtendedRights|ForceChangePassword|GenericAll|GenericWrite|WriteDacl|WriteOwner*1..]->(g2))
                        RETURN LENGTH(p)
                        """

        session = self.driver.session()
        count = 0
        for result in session.run(count_query, domain=self.domain):
            count = result[0]

        session.close()
        self.write_single_cell(
            sheet, 2, 3, "Shortest ACL Path Length: {}".format(count))

    def shortest_derivative_path_auth_users(self, sheet):
        count_query = """MATCH (g1:Group {domain:{domain}})
                        WHERE g1.objectsid = 'S-1-5-11'
                        MATCH (g2:Group {domain:{domain}})
                        WHERE g2.objectsid ENDS WITH "-512"
                        MATCH p = shortestPath((g1)-[:AdminTo|HasSession|MemberOf*1..]->(g2))
                        RETURN LENGTH(p)
                        """

        session = self.driver.session()
        count = 0
        for result in session.run(count_query, domain=self.domain):
            count = result[0]

        session.close()
        self.write_single_cell(
            sheet, 3, 3, "Shortest Derivative Path Length: {}".format(count))

    def shortest_hybrid_path_auth_users(self, sheet):
        count_query = """MATCH (g1:Group {domain:{domain}})
                        WHERE g1.objectsid = 'S-1-5-11'
                        MATCH (g2:Group {domain:{domain}})
                        WHERE g2.objectsid ENDS WITH "-512"
                        MATCH p = shortestPath((g1)-[r*1..]->(g2))
                        WHERE NONE(rel in r WHERE type(rel)="GetChanges")
                        WITH *
                        WHERE NONE(rel in r WHERE type(rel)="GetChangesAll")
                        RETURN LENGTH(p)
                        """

        session = self.driver.session()
        count = 0
        for result in session.run(count_query, domain=self.domain):
            count = result[0]

        session.close()
        self.write_single_cell(
            sheet, 4, 3, "Shortest Hybrid Path Length: {}".format(count))

    def kerberoastable_path_len(self, sheet):
        list_query = """MATCH (u:User {domain:{domain},hasspn:true})
                        MATCH (g:Group {domain:{domain}})
                        WHERE g.objectsid ENDS WITH "-512" AND NOT u.name STARTS WITH "KRBTGT@"
                        MATCH p = shortestPath((u)-[*1..]->(g))
                        RETURN u.name,LENGTH(p)
                        ORDER BY LENGTH(p) ASC
                        """

        session = self.driver.session()
        results = []
        for result in session.run(list_query, domain=self.domain):
            results.append(
                "{} - {}".format(result[0], result[1]))

        session.close()
        self.write_column_data(
            sheet, "Kerberoastable User to DA Path Length", results)

    def asreproastable_path_len(self, sheet):
        list_query = """MATCH (u:User {domain:{domain},dontreqpreauth:True})
                        MATCH (g:Group {domain:{domain}})
                        WHERE g.objectsid ENDS WITH "-512" AND NOT u.name STARTS WITH "KRBTGT@"
                        MATCH p = shortestPath((u)-[*1..]->(g))
                        RETURN u.name,LENGTH(p)
                        ORDER BY LENGTH(p) ASC
                        """

        session = self.driver.session()
        results = []
        for result in session.run(list_query, domain=self.domain):
            results.append(
                "{} - {}".format(result[0], result[1]))

        session.close()
        self.write_column_data(
            sheet, "ASReproastable User to DA Path Length", results)

    def high_admin_comps(self, sheet):
        list_query = """MATCH (c:Computer {domain:{domain}})
                        OPTIONAL MATCH (n)-[:AdminTo]->(c)
                        OPTIONAL MATCH (m)-[:MemberOf*1..]->(:Group)-[:AdminTo]->(c)
                        WITH COLLECT(n) + COLLECT(m) as tempVar,c
                        UNWIND tempVar as admins
                        RETURN c.name,COUNT(DISTINCT(admins))
                        ORDER BY COUNT(DISTINCT(admins)) DESC
                        """

        session = self.driver.session()
        results = []
        for result in session.run(list_query, domain=self.domain):
            count = result[1]
            if (count > 1000):
                results.append(
                    "{} - {}".format(result[0], count))

        session.close()
        self.write_column_data(
            sheet, "Computers with > 1000 Admins: {}", results)

    def server_2003(self, sheet):
        list_query = """MATCH (n:Computer {enabled: True,domain:{domain}})
                            WHERE n.operatingsystem =~ "Windows Server 2003.*"
                            RETURN n.name
                            """
        session = self.driver.session()
        results = []

        for result in session.run(list_query, domain=self.domain):
            results.append(result[0])

        session.close()
        self.write_column_data(
            sheet, "Windows Server 2003: {}", results)

    def server_2008(self, sheet):
        list_query = """MATCH (n:Computer {enabled: True,domain:{domain}})
                            WHERE n.operatingsystem =~ "Windows Server 2008.*"
                            RETURN n.name
                            """
        session = self.driver.session()
        results = []

        for result in session.run(list_query, domain=self.domain):
            results.append(result[0])

        session.close()
        self.write_column_data(
            sheet, "Windows Server 2008(Expire 2020): {}", results)

    def spoolbug(self, sheet):
        list_query = """OPTIONAL MATCH (c1:Computer)-[r1:AdminTo*1..]->(c2:Computer {domain:{domain}})
                            OPTIONAL MATCH (c3:Computer)-[:MemberOf*1..]->(:Group)-[r2:AdminTo]->(c4:Computer {domain:{domain}})
                            WITH collect(c1) + collect(c3) AS temp
                            UNWIND temp as computers
                            RETURN DISTINCT(computers.name),COUNT(computers)
                            ORDER BY computers.name ASC
                            """

        session = self.driver.session()
        results = []

        for result in session.run(list_query, domain=self.domain):
            results.append("{} - {}".format(result[0], result[1]))

        session.close()
        self.write_column_data(
            sheet, "SpoolBug Relay From: {}", results)

    def passnotreqd(self, sheet):
        list_query = """MATCH (u:User)
                            WHERE u.passwordnotreqd = True
                            RETURN u.name,u.enabled
                            """

        session = self.driver.session()
        results = []

        for result in session.run(list_query, domain=self.domain):
            results.append("{} - {}".format(result[0], result[1]))

        session.close()
        self.write_column_data(
            sheet, "PasswordNotReqd: {}", results)

class CriticalAssets(object):
    def __init__(self, driver, domain, workbook):
        self.driver = driver
        self.domain = domain
        self.col_count = 1
        self.workbook = workbook

    def write_column_data(self, sheet, title, results):
        count = len(results)
        offset = 1
        font = styles.Font(bold=True)
        c = sheet.cell(offset, self.col_count)
        c.font = font
        sheet.cell(offset, self.col_count, value=title.format(count))
        for i in xrange(0, count):
            sheet.cell(i + offset + 1, self.col_count, value=results[i])
        self.col_count += 1

    def do_critical_asset_analysis(self):
        func_list = [
            self.admins_on_dc, self.rdp_on_dc, self.gpo_on_dc, self.admin_on_exch,
            self.rdp_on_exch, self.gpo_on_exch, self.da_controllers,
            self.da_sessions, self.gpo_on_da, self.da_equiv_controllers,
            self.da_equiv_sessions, self.gpo_on_da_equiv, self.dcsync
        ]
        sheet = self.workbook._sheets[1]
        for f in func_list:
            s = timer()
            f(sheet)
            print "{} completed in {}s".format(f.__name__, timer() - s)

    def admins_on_dc(self, sheet):
        list_query = """MATCH (g:Group {domain:{domain}})
                    WHERE g.objectsid ENDS WITH "-516"
                    MATCH (c:Computer)-[:MemberOf*1..]->(g)
                    OPTIONAL MATCH (n)-[:AdminTo]->(c)
                    OPTIONAL MATCH (m)-[:MemberOf*1..]->(:Group)-[:AdminTo]->(c)
                    WHERE (n:User OR n:Computer) AND (m:User OR m:Computer)
                    WITH COLLECT(n) + COLLECT(m) as tempVar1
                    UNWIND tempVar1 as tempVar2
                    WITH DISTINCT(tempVar2) as tempVar3
                    RETURN tempVar3.name
                    ORDER BY tempVar3.name ASC"""

        session = self.driver.session()
        results = []

        for result in session.run(list_query, domain=self.domain):
            results.append(result[0])

        session.close()
        self.write_column_data(
            sheet, "Admins on Domain Controllers: {}", results)

    def rdp_on_dc(self, sheet):
        list_query = """MATCH (g:Group {domain:{domain}})
                    WHERE g.objectsid ENDS WITH "-516"
                    MATCH (c:Computer)-[:MemberOf*1..]->(g)
                    OPTIONAL MATCH (n)-[:CanRDP]->(c)
                    OPTIONAL MATCH (m)-[:MemberOf*1..]->(:Group)-[:CanRDP]->(c)
                    WHERE (n:User OR n:Computer) AND (m:User OR m:Computer)

                    WITH COLLECT(n) + COLLECT(m) as tempVar1
                    UNWIND tempVar1 as tempVar2
                    WITH DISTINCT(tempVar2) as tempVar3
                    RETURN tempVar3.name
                    ORDER BY tempVar3.name ASC
                    """
        session = self.driver.session()
        results = []

        for result in session.run(list_query, domain=self.domain):
            results.append(result[0])

        session.close()
        self.write_column_data(
            sheet, "RDPers on Domain Controllers: {}", results)

    def gpo_on_dc(self, sheet):
        list_query = """MATCH (g:Group {domain:{domain}})
                        WHERE g.objectsid ENDS WITH "-516"
                        MATCH (c:Computer)-[:MemberOf*1..]->(g)
                        OPTIONAL MATCH p1 = (g1:GPO)-[r1:GpLink {enforced:true}]->(container1)-[r2:Contains*1..]->(c)
                        OPTIONAL MATCH p2 = (g2:GPO)-[r3:GpLink {enforced:false}]->(container2)-[r4:Contains*1..]->(c)
                        WHERE NONE (x in NODES(p2) WHERE x.blocksinheritance = true AND x:OU AND NOT (g2)-->(x))
                        WITH COLLECT(g1) + COLLECT(g2) AS tempVar1
                        UNWIND tempVar1 as tempVar2
                        WITH DISTINCT(tempVar2) as GPOs
                        OPTIONAL MATCH (n)-[{isacl:true}]->(GPOs)
                        OPTIONAL MATCH (m)-[:MemberOf*1..]->(:Group)-[{isacl:true}]->(GPOs)
                        WITH COLLECT(n) + COLLECT(m) as tempVar1
                        UNWIND tempVar1 as tempVar2
                        RETURN DISTINCT(tempVar2.name)
                        ORDER BY tempVar2.name ASC
                        """

        session = self.driver.session()
        results = []

        for result in session.run(list_query, domain=self.domain):
            results.append(result[0])

        session.close()
        self.write_column_data(
            sheet, "Domain Controller GPO Controllers: {}", results)

    def admin_on_exch(self, sheet):
        list_query = """MATCH (n:Computer)
                        UNWIND n.serviceprincipalnames AS spn
                        MATCH (n) WHERE TOUPPER(spn) CONTAINS "EXCHANGEMDB"
                        WITH n as c
                        MATCH (c)-[:MemberOf*1..]->(g:Group {domain:{domain}})
                        WHERE g.name CONTAINS "EXCHANGE"
                        OPTIONAL MATCH (n)-[:AdminTo]->(c)
                        OPTIONAL MATCH (m)-[:MemberOf*1..]->(:Group)-[:AdminTo]->(c)
                        WITH COLLECT(n) + COLLECT(m) as tempVar1
                        UNWIND tempVar1 AS exchangeAdmins
                        RETURN DISTINCT(exchangeAdmins.name)
                        """

        session = self.driver.session()
        results = []

        for result in session.run(list_query, domain=self.domain):
            results.append(result[0])

        session.close()
        self.write_column_data(
            sheet, "Admins on Exchange Servers: {}", results)

    def rdp_on_exch(self, sheet):
        list_query = """MATCH (n:Computer)
                        UNWIND n.serviceprincipalnames AS spn
                        MATCH (n) WHERE TOUPPER(spn) CONTAINS "EXCHANGEMDB"
                        WITH n as c
                        MATCH (c)-[:MemberOf*1..]->(g:Group {domain:{domain}})
                        WHERE g.name CONTAINS "EXCHANGE"
                        OPTIONAL MATCH (n)-[:CanRDP]->(c)
                        OPTIONAL MATCH (m)-[:MemberOf*1..]->(:Group)-[:CanRDP]->(c)
                        WITH COLLECT(n) + COLLECT(m) as tempVar1
                        UNWIND tempVar1 AS exchangeAdmins
                        RETURN DISTINCT(exchangeAdmins.name)
                        """

        session = self.driver.session()
        results = []

        for result in session.run(list_query, domain=self.domain):
            results.append(result[0])

        session.close()
        self.write_column_data(
            sheet, "RDPers on Exchange Servers: {}", results)

    def gpo_on_exch(self, sheet):
        list_query = """MATCH (n:Computer)
                        UNWIND n.serviceprincipalnames AS spn
                        MATCH (n) WHERE TOUPPER(spn) CONTAINS "EXCHANGEMDB"
                        WITH n as c
                        MATCH (c)-[:MemberOf*1..]->(g:Group {domain:{domain}})
                        WHERE g.name CONTAINS "EXCHANGE"
                        OPTIONAL MATCH p1 = (g1:GPO)-[r1:GpLink {enforced:true}]->(container1)-[r2:Contains*1..]->(c)
                        OPTIONAL MATCH p2 = (g2:GPO)-[r3:GpLink {enforced:false}]->(container2)-[r4:Contains*1..]->(c)
                        WHERE NONE (x in NODES(p2) WHERE x.blocksinheritance = true AND x:OU AND NOT (g2)-->(x))
                        WITH COLLECT(g1) + COLLECT(g2) AS tempVar1
                        UNWIND tempVar1 as tempVar2
                        WITH DISTINCT(tempVar2) as GPOs
                        OPTIONAL MATCH (n)-[{isacl:true}]->(GPOs)
                        OPTIONAL MATCH (m)-[:MemberOf*1..]->(:Group)-[{isacl:true}]->(GPOs)
                        WITH COLLECT(n) + COLLECT(m) as tempVar1
                        UNWIND tempVar1 as tempVar2
                        RETURN DISTINCT(tempVar2.name)
                        ORDER BY tempVar2.name ASC
                        """

        session = self.driver.session()
        results = []

        for result in session.run(list_query, domain=self.domain):
            results.append(result[0])

        session.close()
        self.write_column_data(
            sheet, "Exchange Server GPO Controllers: {}", results)

    def da_controllers(self, sheet):
        list_query = """MATCH (DAUser)-[:MemberOf*1..]->(g:Group {domain:{domain}})
                        WHERE g.objectsid ENDS WITH "-512"
                        OPTIONAL MATCH (n)-[{isacl:true}]->(DAUser)
                        OPTIONAL MATCH (m)-[:MemberOf*1..]->(:Group)-[{isacl:true}]->(DAUser)
                        WITH COLLECT(n) + COLLECT(m) as tempVar1
                        UNWIND tempVar1 AS DAControllers
                        RETURN DISTINCT(DAControllers.name)
                        ORDER BY DAControllers.name ASC
                        """

        session = self.driver.session()
        results = []

        for result in session.run(list_query, domain=self.domain):
            results.append(result[0])

        session.close()
        self.write_column_data(
            sheet, "Domain Admin Controllers: {}", results)

    def da_sessions(self, sheet):
        list_query = """MATCH (c:Computer)-[:HasSession]->()-[:MemberOf*1..]->(g:Group {domain:{domain}})
                        WHERE g.objectsid ENDS WITH "-512"
                        RETURN DISTINCT(c.name)
                        ORDER BY c.name ASC
                        """

        session = self.driver.session()
        results = []

        for result in session.run(list_query, domain=self.domain):
            results.append(result[0])

        session.close()
        self.write_column_data(
            sheet, "Computers with DA Sessions: {}", results)

    def gpo_on_da(self, sheet):
        list_query = """MATCH (DAUser)-[:MemberOf*1..]->(g:Group {domain:{domain}})
                        WHERE g.objectsid ENDS WITH "-512"
                        OPTIONAL MATCH p1 = (g1:GPO)-[r1:GpLink {enforced:true}]->(container1)-[r2:Contains*1..]->(DAUser)
                        OPTIONAL MATCH p2 = (g2:GPO)-[r3:GpLink {enforced:false}]->(container2)-[r4:Contains*1..]->(DAUser)
                        WHERE NONE (x in NODES(p2) WHERE x.blocksinheritance = true AND x:OU AND NOT (g2)-->(x))
                        WITH COLLECT(g1) + COLLECT(g2) AS tempVar1
                        UNWIND tempVar1 as tempVar2
                        WITH DISTINCT(tempVar2) as GPOs
                        OPTIONAL MATCH (n)-[{isacl:true}]->(GPOs)
                        OPTIONAL MATCH (m)-[:MemberOf*1..]->(:Group)-[{isacl:true}]->(GPOs)
                        WITH COLLECT(n) + COLLECT(m) as tempVar1
                        UNWIND tempVar1 as tempVar2
                        RETURN DISTINCT(tempVar2.name)
                        """

        session = self.driver.session()
        results = []

        for result in session.run(list_query, domain=self.domain):
            results.append(result[0])

        session.close()
        self.write_column_data(
            sheet, "Domain Admin GPO Controllers: {}", results)

    def da_equiv_controllers(self, sheet):
        list_query = """MATCH (u:User)-[:MemberOf*1..]->(g:Group {domain:{domain},highvalue:true})
                        OPTIONAL MATCH (n)-[{isacl:true}]->(u)
                        OPTIONAL MATCH (m)-[:MemberOf*1..]->(:Group)-[{isacl:true}]->(u)
                        WITH COLLECT(n) + COLLECT(m) as tempVar
                        UNWIND tempVar as highValueControllers
                        RETURN DISTINCT(highValueControllers.name)
                        ORDER BY highValueControllers.name ASC
                        """

        session = self.driver.session()
        results = []

        for result in session.run(list_query, domain=self.domain):
            results.append(result[0])

        session.close()
        self.write_column_data(
            sheet, "High Value Object Controllers: {}", results)

    def da_equiv_sessions(self, sheet):
        list_query = """MATCH (c:Computer)-[:HasSession]->(u:User)-[:MemberOf*1..]->(g:Group {domain:{domain},highvalue:true})
                        RETURN DISTINCT(c.name)
                        ORDER BY c.name ASC
                        """

        session = self.driver.session()
        results = []

        for result in session.run(list_query, domain=self.domain):
            results.append(result[0])

        session.close()
        self.write_column_data(
            sheet, "High Value User Sessions: {}", results)

    def gpo_on_da_equiv(self, sheet):
        list_query = """MATCH (u:User)-[:MemberOf*1..]->(g:Group {domain:{domain},highvalue:true})
                        OPTIONAL MATCH p1 = (g1:GPO)-[r1:GpLink {enforced:true}]->(container1)-[r2:Contains*1..]->(u)
                        OPTIONAL MATCH p2 = (g2:GPO)-[r3:GpLink {enforced:false}]->(container2)-[r4:Contains*1..]->(u)
                        WHERE NONE (x in NODES(p2) WHERE x.blocksinheritance = true AND x:OU AND NOT (g2)-->(x))
                        WITH COLLECT(g1) + COLLECT(g2) AS tempVar1
                        UNWIND tempVar1 as tempVar2
                        WITH DISTINCT(tempVar2) as GPOs
                        OPTIONAL MATCH (n)-[{isacl:true}]->(GPOs)
                        OPTIONAL MATCH (m)-[:MemberOf*1..]->(:Group)-[{isacl:true}]->(GPOs)
                        WITH COLLECT(n) + COLLECT(m) as tempVar1
                        UNWIND tempVar1 as tempVar2
                        RETURN DISTINCT(tempVar2.name)
                        ORDER BY tempVar2.name ASC
                        """

        session = self.driver.session()
        results = []

        for result in session.run(list_query, domain=self.domain):
            results.append(result[0])

        session.close()
        self.write_column_data(
            sheet, "High Value User GPO Controllers: {}", results)

    def dcsync(self, sheet):
        list_query = """MATCH (n1)-[:MemberOf|GetChanges*1..]->(u:Domain {name: {domain}}) 
                            WITH n1,u 
                            MATCH (n1)-[:MemberOf|GetChangesAll*1..]->(u) 
                            WITH n1,u 
                            MATCH (n1)-[:MemberOf|GetChanges|GetChangesAll*1..]->(u) 
                            RETURN DISTINCT(n1.name)
                            """
        session = self.driver.session()
        results = []

        for result in session.run(list_query, domain=self.domain):
            results.append(result[0])

        session.close()
        self.write_column_data(
            sheet, "DCSync Principals: {}", results)

class CrossDomain(object):
    def __init__(self, driver, domain, workbook):
        self.driver = driver
        self.domain = domain
        self.col_count = 1
        self.workbook = workbook

    def write_column_data(self, sheet, title, results):
        count = len(results)
        offset = 1
        font = styles.Font(bold=True)
        c = sheet.cell(offset, self.col_count)
        c.font = font
        sheet.cell(offset, self.col_count, value=title.format(count))
        for i in xrange(0, count):
            sheet.cell(i + offset + 1, self.col_count, value=results[i])
        self.col_count += 1

    def write_single_cell(self, sheet, row, column, text):
        sheet.cell(row, column, value=text)

    def do_cross_domain_analysis(self):
        func_list = [
            self.foreign_admins, self.foreign_gpo_controllers,
            self.foreign_user_controllers, self.foreign_session
        ]
        sheet = self.workbook._sheets[3]

        for f in func_list:
            s = timer()
            f(sheet)
            print "{} completed in {}s".format(f.__name__, timer() - s)

    def foreign_admins(self, sheet):
        list_query = """MATCH (c:Computer {domain:{domain}})
                        OPTIONAL MATCH (n)-[:AdminTo]->(c)
                        WHERE (n:User OR n:Computer) AND NOT n.domain = c.domain
                        OPTIONAL MATCH (m)-[:MemberOf*1..]->(:Group)-[:AdminTo]->(c)
                        WHERE (m:User OR m:Computer) AND NOT m.domain = c.domain
                        WITH COLLECT(n) + COLLECT(m) AS tempVar,c
                        UNWIND tempVar AS foreignAdmins
                        RETURN c.name,COUNT(DISTINCT(foreignAdmins))
                        ORDER BY COUNT(DISTINCT(foreignAdmins)) DESC
                        """

        session = self.driver.session()
        results = []

        for result in session.run(list_query, domain=self.domain):
            results.append("{} - {}".format(result[0], result[1]))

        session.close()
        self.write_column_data(
            sheet, "Computers with Foreign Admins: {}", results)

    def foreign_gpo_controllers(self, sheet):
        list_query = """MATCH (g:GPO)
                        WHERE SPLIT(g.name,'@')[1] = {domain}
                        OPTIONAL MATCH (n)-[{isacl:true}]->(g)
                        WHERE (n:User OR n:Computer) AND NOT n.domain = {domain}
                        OPTIONAL MATCH (m)-[:MemberOf*1..]->(:Group)-[{isacl:true}]->(g)
                        WHERE (m:User OR m:Computer) AND NOT m.domain = {domain}
                        WITH COLLECT(n) + COLLECT(m) AS tempVar,g
                        UNWIND tempVar AS foreignGPOControllers
                        RETURN g.name,COUNT(DISTINCT(foreignGPOControllers))
                        ORDER BY COUNT(DISTINCT(foreignGPOControllers)) DESC
                        """

        session = self.driver.session()
        results = []

        for result in session.run(list_query, domain=self.domain):
            results.append("{} - {}".format(result[0], result[1]))

        session.close()
        self.write_column_data(
            sheet, "GPOs with Foreign Controllers: {}", results)

    def foreign_user_controllers(self, sheet):
        list_query = """MATCH (g:Group {domain:{domain}})
                        OPTIONAL MATCH (n)-[{isacl:true}]->(g)
                        WHERE (n:User OR n:Computer) AND NOT n.domain = g.domain
                        OPTIONAL MATCH (m)-[:MemberOf*1..]->(:Group)-[{isacl:true}]->(g)
                        WHERE (m:User OR m:Computer) AND NOT m.domain = g.domain
                        WITH COLLECT(n) + COLLECT(m) AS tempVar,g
                        UNWIND tempVar AS foreignGroupControllers
                        RETURN g.name,COUNT(DISTINCT(foreignGroupControllers))
                        ORDER BY COUNT(DISTINCT(foreignGroupControllers)) DESC
                        """

        session = self.driver.session()
        results = []

        for result in session.run(list_query, domain=self.domain):
            results.append("{} - {}".format(result[0], result[1]))

        session.close()
        self.write_column_data(
            sheet, "Groups with Foreign Controllers: {}", results)

    def foreign_session(self, sheet):
        list_query = """MATCH (s:Computer {domain:{domain}})-[r:HasSession*1]->(t:User)
                        WHERE NOT s.domain = t.domain
                        RETURN s.name, COUNT(t) as count
                        ORDER BY count DESC
                        """

        session = self.driver.session()
        results = []

        for result in session.run(list_query, domain=self.domain):
            results.append("{} - {}".format(result[0], result[1]))

        session.close()
        self.write_column_data(
            sheet, "Computers with Foreign Session: {}", results)

class Organization(object):
    def __init__(self, driver, domain, workbook):
        self.driver = driver
        self.domain = domain
        self.col_count = 1
        self.workbook = workbook

    def write_column_data(self, sheet, title, results):
        count = len(results)
        offset = 1
        font = styles.Font(bold=True)
        c = sheet.cell(offset, self.col_count)
        c.font = font
        sheet.cell(offset, self.col_count, value=title.format(count))
        for i in xrange(0, count):
            sheet.cell(i + offset + 1, self.col_count, value=results[i])
        self.col_count += 1

    def write_single_cell(self, sheet, row, column, text):
        sheet.cell(row, column, value=text)

    def do_Organization(self):
        func_list = [
            self.da_members, self.high_value_members, self.user_local_admin,
            self.groups_local_admin, self.local_admin_inbound,
            self.most_session_user, self.most_session_computer, self.laps,
            self.highvalue_protectedgroup
        ]
        sheet = self.workbook._sheets[4]

        for f in func_list:
            s = timer()
            f(sheet)
            print "{} completed in {}s".format(f.__name__, timer() - s)

    def da_members(self, sheet):
        list_query = """MATCH (n:Group {domain:{domain}})
                            WHERE n.objectsid =~ "(?i)S-1-5.*-512" WITH n
                            MATCH (n)<-[r:MemberOf*1..]-(m)
                            RETURN m.name
                            ORDER BY m.name ASC
                            """
        session = self.driver.session()
        results = []

        for result in session.run(list_query, domain=self.domain):
            results.append(result[0])

        session.close()
        self.write_column_data(
            sheet, "Domain Admin members: {}", results)

    def high_value_members(self, sheet):
        list_query = """MATCH (n)-[:MemberOf*1..]->(g {highvalue: True,domain: {domain}})
                        RETURN DISTINCT n.name
                        ORDER BY n.name ASC
                        """
        session = self.driver.session()
        results = []

        for result in session.run(list_query, domain=self.domain):
            results.append(result[0])

        session.close()
        self.write_column_data(
            sheet, "High Value Group members: {}", results)

    def user_local_admin(self, sheet):
        list_query = """MATCH (u:User {domain:{domain}})
                            WITH u
                            OPTIONAL MATCH (u:User)-[r:AdminTo]->(c:Computer {domain:{domain}})
                            WITH u,COUNT(c) as expAdmin
                            OPTIONAL MATCH (u:User)-[r:MemberOf*1..]->(g:Group)-[r2:AdminTo]->(c:Computer {domain:{domain}})
                            WHERE NOT (u)-[:AdminTo]->(c)
                            WITH u,expAdmin,COUNT(DISTINCT(c)) as unrolledAdmin
                            RETURN u.name,expAdmin + unrolledAdmin as totalAdmin
                            ORDER BY totalAdmin DESC
                            LIMIT 100
                            """
        session = self.driver.session()
        results = []

        for result in session.run(list_query, domain=self.domain):
            results.append("{} - {}".format(result[0], result[1]))

        session.close()
        self.write_column_data(
            sheet, "Top 100 Users with Local Admin: {}", results)

    def groups_local_admin(self, sheet):
        list_query = """MATCH (u:Group {domain:{domain}})
                            WITH u
                            OPTIONAL MATCH (u:Group)-[r:AdminTo]->(c:Computer {domain:{domain}})
                            WITH u,COUNT(c) as expAdmin
                            OPTIONAL MATCH (u:Group)-[r:MemberOf*1..]->(g:Group)-[r2:AdminTo]->(c:Computer {domain:{domain}})
                            WHERE NOT (u)-[:AdminTo]->(c)
                            WITH u,expAdmin,COUNT(DISTINCT(c)) as unrolledAdmin
                            RETURN u.name,expAdmin + unrolledAdmin as totalAdmin
                            ORDER BY totalAdmin DESC
                            LIMIT 100
                            """
        session = self.driver.session()
        results = []

        for result in session.run(list_query, domain=self.domain):
            results.append("{} - {}".format(result[0], result[1]))

        session.close()
        self.write_column_data(
            sheet, "Top 100 Groups with Local Admin: {}", results)

    def local_admin_inbound(self, sheet):
        list_query = """MATCH (c:Computer {domain:{domain}})
                            OPTIONAL MATCH (u1)-[:AdminTo]->(c)
                            OPTIONAL MATCH (u2)-[:MemberOf*1..]->(:Group)-[:AdminTo]->(c)
                            WITH COLLECT(u1) + COLLECT(u2) as tempVar,c
                            UNWIND tempVar as admins
                            WITH c,COUNT(DISTINCT(admins)) as adminCount
                            RETURN c.name,adminCount
                            ORDER BY adminCount DESC
                            LIMIT 100
                            """
        session = self.driver.session()
        results = []

        for result in session.run(list_query, domain=self.domain):
            results.append("{} - {}".format(result[0], result[1]))

        session.close()
        self.write_column_data(
            sheet, "Top 100 Local Admin Inbound: {}", results)

    def most_session_computer(self, sheet):
        list_query = """Match (c:Computer {domain:{domain}})-[r:HasSession]->(u:User)
                            WITH c,COUNT(u) as session
                            RETURN c.name,session
                            ORDER BY session DESC
                            LIMIT 100
                            """
        session = self.driver.session()
        results = []

        for result in session.run(list_query, domain=self.domain):
            results.append("{} - {}".format(result[0], result[1]))

        session.close()
        self.write_column_data(
            sheet, "Top 100 Computers with most sessions: {}", results)

    def most_session_user(self, sheet):
        list_query = """MATCH (u:User)<-[s:HasSession]-(c:Computer {domain:{domain}})
                            WITH u, count(s) as session
                            RETURN u.name,session
                            ORDER BY session DESC
                            LIMIT 100
                            """
        session = self.driver.session()
        results = []

        for result in session.run(list_query, domain=self.domain):
            results.append("{} - {}".format(result[0], result[1]))

        session.close()
        self.write_column_data(
            sheet, "Top 100 Users with most sessions: {}", results)

    def laps(self, sheet):
        list_query = """MATCH(n)-[:ReadLAPSPassword*1..]->(c:Computer {domain:{domain}})
                        RETURN n.name, count(c) as count
                        ORDER BY count DESC
                        """
        session = self.driver.session()
        results = []

        for result in session.run(list_query, domain=self.domain):
            results.append("{} - {}".format(result[0], result[1]))

        session.close()
        self.write_column_data(
            sheet, "LAPS Rights: {}", results)

    def highvalue_protectedgroup(self, sheet):
        list_query = """MATCH (g:Group {domain:{domain}})
                        WHERE g.objectsid ENDS WITH '-525'
                        OPTIONAL MATCH (u:User {domain:{domain}})-[:MemberOf*1..]->(:Group {highvalue:True})
                        WHERE NOT (u)-[:MemberOf*1..]->(g)
                        RETURN DISTINCT(u.name)
                        """
        session = self.driver.session()
        results = []

        for result in session.run(list_query, domain=self.domain):
            results.append(result[0])

        session.close()
        self.write_column_data(
            sheet, "High Value Users not in ProtectedGroup: {}", results)

class kerberos(object):
    def __init__(self, driver, domain, workbook):
        self.driver = driver
        self.domain = domain
        self.col_count = 1
        self.workbook = workbook

    def write_column_data(self, sheet, title, results):
        count = len(results)
        offset = 4
        font = styles.Font(bold=True)
        c = sheet.cell(offset, self.col_count)
        c.font = font
        sheet.cell(offset, self.col_count, value=title.format(count))
        for i in xrange(0, count):
            sheet.cell(i + offset + 1, self.col_count, value=results[i])
        self.col_count += 1

    def write_single_cell(self, sheet, row, column, text):
        sheet.cell(row, column, value=text)

    def do_kerberos(self):
        func_list = [
            self.shortest_path_everyone, self.shortest_path_auth_users,
            self.shortest_path_domain_users, self.unconstrained,
            self.allow_to_delegation, self.kerberoastable,self.kerberoast_to_highvalue,
            self.asreproastable, self.asrep_to_highvalue
        ]
        sheet = self.workbook._sheets[5]
        self.write_single_cell(
            sheet, 1, 1, "Domain Users to Unconstrained System")
        self.write_single_cell(sheet, 1, 2, "Everyone to Unconstrained System")
        self.write_single_cell(
            sheet, 1, 3, "Authenticated Users to Unconstrained System")
        font = styles.Font(bold=True)
        sheet.cell(1, 1).font = font
        sheet.cell(1, 2).font = font
        sheet.cell(1, 3).font = font

        for f in func_list:
            s = timer()
            f(sheet)
            print "{} completed in {}s".format(f.__name__, timer() - s)

    def shortest_path_domain_users(self, sheet):
        count_query = """MATCH (g1:Group {domain:{domain}})
                        WHERE g1.objectsid ENDS WITH "-513"
                        MATCH (c:Computer {domain:{domain}})
                        WHERE c.unconstraineddelegation = True
                        MATCH p = shortestPath((g1)-[r:MemberOf|HasSession|AdminTo|AllExtendedRights|AddMember|ForceChangePassword|GenericAll|GenericWrite|Owns|WriteDacl|WriteOwner|CanRDP|ExecuteDCOM|AllowedToDelegate|ReadLAPSPassword|Contains|GpLink|AddAllowedToAct|AllowedToAct*1..]->(c))
                        RETURN LENGTH(p)
                        """
        session = self.driver.session()
        count = 0
        for result in session.run(count_query, domain=self.domain):
            count = result[0]

        session.close()
        self.write_single_cell(
            sheet, 2, 1, "Shortest Path Length: {}".format(count))

    def shortest_path_everyone(self, sheet):
        count_query = """MATCH (g1:Group {domain:{domain}})
                        WHERE g1.objectsid = 'S-1-1-0'
                        MATCH (c:Computer {domain:{domain}})
                        WHERE c.unconstraineddelegation = True
                        MATCH p = shortestPath((g1)-[r:MemberOf|HasSession|AdminTo|AllExtendedRights|AddMember|ForceChangePassword|GenericAll|GenericWrite|Owns|WriteDacl|WriteOwner|CanRDP|ExecuteDCOM|AllowedToDelegate|ReadLAPSPassword|Contains|GpLink|AddAllowedToAct|AllowedToAct*1..]->(c))
                        RETURN LENGTH(p)
                        """

        session = self.driver.session()
        count = 0
        for result in session.run(count_query, domain=self.domain):
            count = result[0]

        session.close()
        self.write_single_cell(
            sheet, 2, 2, "Shortest Path Length: {}".format(count))

    def shortest_path_auth_users(self, sheet):
        count_query = """MATCH (g1:Group {domain:{domain}})
                        WHERE g1.objectsid = 'S-1-5-11'
                        MATCH (c:Computer {domain:{domain}})
                        WHERE c.unconstraineddelegation = True
                        MATCH p = shortestPath((g1)-[r:MemberOf|HasSession|AdminTo|AllExtendedRights|AddMember|ForceChangePassword|GenericAll|GenericWrite|Owns|WriteDacl|WriteOwner|CanRDP|ExecuteDCOM|AllowedToDelegate|ReadLAPSPassword|Contains|GpLink|AddAllowedToAct|AllowedToAct*1..]->(c))
                        RETURN LENGTH(p)
                        """
        session = self.driver.session()
        count = 0
        for result in session.run(count_query, domain=self.domain):
            count = result[0]

        session.close()
        self.write_single_cell(
            sheet, 2, 3, "Shortest Path Length: {}".format(count))

    def unconstrained(self, sheet):
        list_query = """MATCH (g:Group {domain: {domain}})
                            WHERE g.objectsid =~ "(?i)S-1-5.*-516"
                            MATCH (c:Computer {unconstraineddelegation: True,domain: {domain}})
                            WHERE NOT (c)-[:MemberOf]->(g)
                            RETURN c.name
                            """
        session = self.driver.session()
        results = []

        for result in session.run(list_query, domain=self.domain):
            results.append(result[0])

        session.close()
        self.write_column_data(
            sheet, "Unconstrained systems: {}", results)

    def allow_to_delegation(self, sheet):
        list_query = """MATCH (n:User {domain: {domain}})
                            WHERE n.sensitive = false
                            MATCH (n)-[r:MemberOf*1..]->(g:Group {highvalue: True,domain: {domain}})
                            RETURN DISTINCT n.name
                            ORDER BY n.name ASC
                            """
        session = self.driver.session()
        results = []

        for result in session.run(list_query, domain=self.domain):
            results.append(result[0])

        session.close()
        self.write_column_data(
            sheet, "Sensitive High Value Users: {}", results)

    def kerberoastable(self, sheet):
        list_query = """MATCH (u:User {domain:{domain},hasspn:true})
                        WHERE NOT u.name STARTS WITH "KRBTGT@"
                        RETURN u.name
                        ORDER BY u.name DESC
                        """

        session = self.driver.session()
        results = []
        for result in session.run(list_query, domain=self.domain):
            results.append(result[0])

        session.close()
        self.write_column_data(
            sheet, "Kerberoastable Users", results)

    def kerberoast_to_highvalue(self, sheet):
        list_query = """MATCH (u:User {hasspn:true}),(n),p=shortestpath((u)-[*1..]->(n))
                        WHERE n.highvalue = true AND n.domain = {domain}
                        RETURN DISTINCT(u.name)
                        ORDER BY u.name ASC
                        """

        session = self.driver.session()
        results = []
        for result in session.run(list_query, domain=self.domain):
            results.append(result[0])

        session.close()
        self.write_column_data(
            sheet, "Kerberoast to highvalue", results)

    def asreproastable(self, sheet):
        list_query = """MATCH (u:User {domain:{domain},dontreqpreauth:True})
                        RETURN u.name
                        ORDER BY u.name DESC
                        """

        session = self.driver.session()
        results = []
        for result in session.run(list_query, domain=self.domain):
            results.append(result[0])

        session.close()
        self.write_column_data(
            sheet, "ASReproastable Users", results)

    def asrep_to_highvalue(self, sheet):
        list_query = """MATCH (u:User {dontreqpreauth:True}),(n),p=shortestpath((u)-[*1..]->(n))
                        WHERE n.highvalue = true AND n.domain = {domain}
                        RETURN DISTINCT(u.name)
                        ORDER BY u.name ASC
                        """

        session = self.driver.session()
        results = []
        for result in session.run(list_query, domain=self.domain):
            results.append(result[0])

        session.close()
        self.write_column_data(
            sheet, "AsReproast to highvalue", results)

class MainMenu(cmd.Cmd):
    def __init__(self):
        self.m = Messages()
        self.url = "bolt://localhost:7687"
        self.username = "neo4j"
        self.password = "neo4jj"
        self.driver = None
        self.connected = False
        self.num_nodes = 500
        self.filename = "BloodHoundAnalytics.xlsx"
        if (len(sys.argv) < 2):
            print "No domain specified."
            print "Usage: python {} DOMAINNAME".format(sys.argv[0])
            sys.exit()
        self.domain = sys.argv[1].upper()
        self.domain_validated = False

        cmd.Cmd.__init__(self)

    def do_changefilename(self, args):
        if args == "":
            print "No filename specified"
            return
        self.filename = args
        print "Change filename to {}".format(self.filename)

    def do_changedomain(self, args):
        if args == "":
            print "No domain specified"
            return
        self.domain_validated = False
        self.domain = args.upper()
        self.validate_domain()

    def cmdloop(self):
        while True:
            self.m.title()
            try:
                cmd.Cmd.cmdloop(self)
            except KeyboardInterrupt:
                if self.driver is not None:
                    self.driver.close()
                raise KeyboardInterrupt

    def do_dbconfig(self, args):
        "Configure connection settings to the neo4j database"
        print "Current Settings:"
        print "DB Url: {}".format(self.url)
        print "DB Username: {}".format(self.username)
        print "DB Password: {}".format(self.password)
        print ""
        self.url = self.m.input_default("Enter DB URL", self.url)
        self.username = self.m.input_default(
            "Enter DB Username", self.username)
        self.password = self.m.input_default(
            "Enter DB Password", self.password)
        print ""
        print "New Settings:"
        print "DB Url: {}".format(self.url)
        print "DB Username: {}".format(self.username)
        print "DB Password: {}".format(self.password)
        print ""
        print "Testing DB Connection"
        self.test_db_conn()

    def do_exit(self, args):
        raise KeyboardInterrupt

    def do_connect(self, args):
        self.test_db_conn()

    def test_db_conn(self):
        self.connected = False
        if self.driver is not None:
            self.driver.close()
        try:
            self.driver = GraphDatabase.driver(
                self.url, auth=(self.username, self.password))
            self.connected = True
            print "Database Connection Successful!"
            self.validate_domain()
        except Exception as e:
            print e
            self.connected = False
            print "Database Connection Failed. Check your settings."

    def validate_domain(self):
        if not self.connected:
            print "Cant validate domain. Connect using connect first"
            return

        print "Validating Selected Domain"
        session = self.driver.session()
        for result in session.run("MATCH (n {domain:{domain}}) RETURN COUNT(n)", domain=self.domain):
            if (int(result[0]) > 0):
                print "Domain {domain} validated!".format(domain=self.domain)
                self.domain_validated = True
                self.create_workbook()
                self.create_analytics()
            else:
                print "Invalid domain specified, use changedomain to pick a new one"

    def do_startanalysis(self, args):
        if not self.connected:
            print "Not connected to database. Use connect command"
            return

        if not self.domain_validated:
            print "Invalid domain or not validated. Use changedomain command"
            return

        print "----------------------------------"
        print "Generating Front Page"
        print "----------------------------------"
        print ""
        self.front.do_front_page_analysis()
        print "----------------------------------"
        print "Generating Critical Assets Page"
        print "----------------------------------"
        print ""
        self.crit.do_critical_asset_analysis()
        print "----------------------------------"
        print "Generating Low Hanging Fruit Page"
        print "----------------------------------"
        print ""
        self.low.do_low_hanging_fruit_analysis()
        print "----------------------------------"
        print "Generating Cross Domain Page"
        print "----------------------------------"
        print ""
        self.cross.do_cross_domain_analysis()
        print "----------------------------------"
        print "Generating High Privileges"
        print "----------------------------------"
        print ""
        self.org.do_Organization()
        print ""
        print "----------------------------------"
        print "Generating Kerberos"
        print "----------------------------------"
        print ""
        self.kerb.do_kerberos()
        print ""
        print "Analytics Complete! Saving workbook to {}".format(self.filename)
        self.save_workbook()

    def create_analytics(self):
        self.crit = CriticalAssets(self.driver, self.domain, self.workbook)
        self.low = LowHangingFruit(self.driver, self.domain, self.workbook)
        self.front = FrontPage(self.driver, self.domain, self.workbook)
        self.cross = CrossDomain(self.driver, self.domain, self.workbook)
        self.org = Organization(self.driver, self.domain, self.workbook)
        self.kerb = kerberos(self.driver, self.domain, self.workbook)
        self.save_workbook()

    def create_workbook(self):
        wb = Workbook()
        ws = wb.active
        ws.title = '{domain} Overview'.format(domain=self.domain)
        wb.create_sheet(title="Critical Assets")
        wb.create_sheet(title="Low Hanging Fruit")
        wb.create_sheet(title="Cross Domain Attacks")
        wb.create_sheet(title="Organization")
        wb.create_sheet(title="Kerberos")
        self.workbook = wb

    def save_workbook(self):
        for worksheet in self.workbook._sheets:
            for col in worksheet.columns:
                max_length = 0
                column = get_column_letter(
                    col[0].column)  # Get the column name
                for cell in col:
                    try:  # Necessary to avoid error on empty cells
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                worksheet.column_dimensions[column].width = adjusted_width
        self.workbook.save(self.filename)

if __name__ == '__main__':
    try:
        MainMenu().cmdloop()
    except KeyboardInterrupt:
        print "Exiting"
        sys.exit()
